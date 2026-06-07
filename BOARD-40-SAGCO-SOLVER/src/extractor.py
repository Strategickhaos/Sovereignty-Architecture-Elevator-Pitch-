"""
BOARD-40-SAGCO-SOLVER — extractor.py
Extract text, rows, and symbols from any supported file type.
Pandas-free: uses csv.DictReader and openpyxl (with fallback).
"""

import csv
import json
import re
from pathlib import Path


def extract_text(path: str) -> str:
    """Read txt/md/yaml/json/py/rs as plain text."""
    p = Path(path)
    try:
        return p.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return f"[read error: {e}]"


def extract_csv(path: str) -> list:
    """Parse CSV using csv.DictReader — no pandas required."""
    rows = []
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))
    return rows


def extract_xlsx(path: str) -> list:
    """Parse XLSX via openpyxl; falls back to CSV if openpyxl is absent."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        rows = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h) if h is not None else f"col{j}" for j, h in enumerate(row)]
                else:
                    rows.append(dict(zip(headers, row)))
        return rows
    except ImportError:
        # fallback: try reading as CSV
        return extract_csv(path)
    except Exception as e:
        return [{"error": str(e)}]


def _extract_symbols(text: str) -> list:
    """Pull out LaTeX/math-like symbols and tokens."""
    pattern = r"[∫∑∏√∞≤≥≠±×÷→←↔∈∉⊂⊃∪∩αβγδεζηθιλμνξπρστυφχψω]"
    return list(set(re.findall(pattern, text)))


def extract(path: str) -> dict:
    """
    Unified extractor. Returns:
    {text, rows, symbols, metadata}
    """
    p = Path(path)
    ext = p.suffix.lower().lstrip(".")
    text = ""
    rows = []

    if ext in ("txt", "md", "yaml", "yml", "json", "py", "rs", "toml", "sh", "ts", "js"):
        text = extract_text(path)
        # also parse json/yaml into rows
        if ext == "json":
            try:
                data = json.loads(text)
                if isinstance(data, list):
                    rows = data
                elif isinstance(data, dict):
                    rows = [data]
            except Exception:
                pass
        elif ext in ("yaml", "yml"):
            try:
                import yaml as _yaml
                data = _yaml.safe_load(text)
                if isinstance(data, list):
                    rows = data
                elif isinstance(data, dict):
                    rows = [data]
            except Exception:
                pass

    elif ext == "csv":
        rows = extract_csv(path)
        text = extract_text(path)

    elif ext in ("xlsx", "xls"):
        rows = extract_xlsx(path)
        text = " ".join(str(v) for row in rows for v in row.values() if v is not None)

    else:
        # generic fallback
        text = extract_text(path)

    symbols = _extract_symbols(text)

    return {
        "text":     text,
        "rows":     rows,
        "symbols":  symbols,
        "metadata": {
            "extension": ext,
            "name":      p.name,
            "char_count": len(text),
            "row_count":  len(rows),
        },
    }
